"""
MAPID BI - Historical Email Scraper
Fetches old emails via IMAP, matches to known clients, exports to Excel.

Usage:
  pip install imapclient openpyxl python-dotenv supabase
  python scripts/email-historical-scraper.py

Reads credentials from .env.local in project root.
"""

import os
import sys
import json
import re
import time
import socket
import email
from email.policy import default
from datetime import datetime, timedelta
from pathlib import Path

# Prevent indefinite hanging on network operations
socket.setdefaulttimeout(60)

# Load .env.local from project root
try:
    from dotenv import load_dotenv
    env_path = Path(__file__).resolve().parent.parent / ".env.local"
    if env_path.exists():
        load_dotenv(env_path)
        print(f"[INFO] Loaded .env.local from {env_path}")
except ImportError:
    pass

import imapclient
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from supabase import create_client

# --- Config ---
IMAP_HOST = os.environ.get("IMAP_HOST", "imap.privateemail.com")
IMAP_PORT = int(os.environ.get("IMAP_PORT", "993"))
IMAP_USER = os.environ.get("IMAP_USER", "")
IMAP_PASS = os.environ.get("IMAP_PASS", "")
LOOKBACK_DAYS = int(os.environ.get("LOOKBACK_DAYS", 1))
SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY", "")


def get_client_names():
    """Fetch active client names directly from Supabase."""
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    projects = supabase.table("kanban_projects").select("client").execute()
    leads = supabase.table("pse_leads").select("lead_name").execute()

    project_clients = [r["client"] for r in (projects.data or []) if r.get("client")]
    lead_names = [r["lead_name"] for r in (leads.data or []) if r.get("lead_name")]
    unique = sorted(set(project_clients + lead_names))
    print(f"[INFO] Fetched {len(unique)} client names from Supabase")
    return unique


def connect_imap():
    """Connect to IMAP server."""
    server = imapclient.IMAPClient(IMAP_HOST, port=IMAP_PORT, ssl=True)
    # Patch the internal imaplib session to support utf-8 encoding for unicode passwords
    server._imap._encoding = "utf-8"
    server.login(IMAP_USER, IMAP_PASS)
    print(f"[INFO] Connected to IMAP: {IMAP_HOST}")
    return server


def search_emails_for_client(server, client_name, since_date):
    """Search emails matching client name in subject."""
    server.select_folder("INBOX", readonly=True)

    since_str = since_date.strftime("%d-%b-%Y")
    all_ids = set()

    # Escape quotes and parens in client name for IMAP
    safe_name = client_name.replace('"', '\\"').replace('(', '\\(').replace(')', '\\)')
    search_str = f'(SUBJECT "{safe_name}" SINCE {since_str})'

    try:
        status, data = server._imap.uid('SEARCH', None, search_str.encode('utf-8'))
        if status == 'OK' and data and data[0]:
            uids = data[0].split()
            all_ids.update(int(uid) for uid in uids if uid.strip())
    except Exception as e:
        print(f"  [WARN] Search failed for '{client_name}': {e}")

    return list(all_ids)


def fetch_emails(server, email_ids, max_emails=100):
    """Fetch email content for given IDs."""
    if not email_ids:
        return []

    email_ids = sorted(email_ids, reverse=True)[:max_emails]

    emails = []
    for eid in email_ids:
        try:
            # Partial fetch: grab up to 20KB (plenty for headers + 5000 chars of body), skips large attachments
            raw = server.fetch([eid], ["BODY.PEEK[]<0.20000>"])[eid]
            raw_email = None
            for key, val in raw.items():
                if isinstance(val, bytes):
                    raw_email = val
                    break
            
            if not raw_email:
                continue

            msg = email.message_from_bytes(raw_email, policy=default)
            subject = str(msg.get("Subject", ""))
            from_addr = str(msg.get("From", ""))
            to_addr = str(msg.get("To", ""))
            date_str = str(msg.get("Date", ""))
            
            body = ""
            try:
                body_part = msg.get_body(preferencelist=('plain', 'html'))
                if body_part:
                    body = body_part.get_content()
                else:
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() in ["text/plain", "text/html"]:
                                payload = part.get_payload(decode=True)
                                if payload:
                                    body = payload.decode(part.get_content_charset() or "utf-8", errors="replace")
                                break
                    else:
                        payload = msg.get_payload(decode=True)
                        if payload:
                            body = payload.decode(msg.get_content_charset() or "utf-8", errors="replace")
            except Exception as e:
                print(f"  [WARN] Failed to parse body for {eid}: {e}")

            body = body[:5000] if body else ""

            emails.append({
                "id": eid,
                "from": from_addr,
                "to": to_addr,
                "subject": subject,
                "date": date_str,
                "body": body,
            })
        except Exception as e:
            print(f"  [WARN] Failed to fetch email {eid}: {e}")

    return emails


def export_to_excel(results, filename="email-data-export.xlsx"):
    """Export all email data to Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Data"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = ["Client Name", "Date", "From", "To", "Subject", "Email Body", "Email Count"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    row = 2
    for client_data in results:
        client_name = client_data["client_name"]
        emails = client_data["emails"]

        if not emails:
            ws.cell(row=row, column=1, value=client_name)
            ws.cell(row=row, column=2, value="No emails found")
            ws.cell(row=row, column=7, value=0)
            row += 1
            continue

        for email in emails:
            ws.cell(row=row, column=1, value=client_name)
            ws.cell(row=row, column=2, value=email["date"])
            ws.cell(row=row, column=3, value=email["from"])
            ws.cell(row=row, column=4, value=email["to"])
            ws.cell(row=row, column=5, value=email["subject"])
            ws.cell(row=row, column=6, value=email["body"])
            ws.cell(row=row, column=7, value=len(emails))
            row += 1

    # Auto-width columns
    column_widths = {
        1: 25,  # Client Name
        2: 25,  # Date
        3: 35,  # From
        4: 35,  # To
        5: 50,  # Subject
        6: 80,  # Body
        7: 12,  # Email Count
    }
    for col, width in column_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(filename)
    print(f"\n[INFO] Excel exported to: {filename}")
    return filename


def upload_to_supabase(results):
    """Upload raw emails to client_emails table in Supabase."""
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    print("\n[INFO] Uploading emails to Supabase...")
    total_uploaded = 0
    
    for client_data in results:
        client_name = client_data["client_name"]
        emails = client_data["emails"]
        
        if not emails:
            continue
            
        rows = []
        for email in emails:
            date_val = email["date"] if email["date"] else None
            
            rows.append({
                "client_name": client_name,
                "subject": email["subject"],
                "from_addr": email["from"],
                "to_addr": email["to"],
                "email_date": date_val,
                "body": email["body"]
            })
            
        try:
            # Bulk insert per client
            supabase.table("client_emails").insert(rows).execute()
            total_uploaded += len(rows)
            print(f"  Uploaded {len(rows)} emails for {client_name}")
        except Exception as e:
            print(f"  [WARN] Failed to upload emails for {client_name}: {e}")
            
    print(f"[INFO] Successfully uploaded {total_uploaded} total emails to Supabase")
    return total_uploaded


def main():
    missing = []
    if not IMAP_USER:
        missing.append("IMAP_USER")
    if not IMAP_PASS:
        missing.append("IMAP_PASS")
    if missing:
        print(f"[ERROR] Missing env: {', '.join(missing)}")
        sys.exit(1)

    print("=" * 60)
    print("MAPID BI - Historical Email Scraper (Excel Export)")
    print(f"Looking back {LOOKBACK_DAYS} days")
    print("=" * 60)

    # 1. Fetch client names
    clients = get_client_names()
    if not clients:
        print("[INFO] No clients found. Exiting.")
        return

    # 2. Connect IMAP
    server = connect_imap()

    # 3. Process each client
    since_date = datetime.now() - timedelta(days=LOOKBACK_DAYS)
    results = []

    for i, client in enumerate(clients):
        print(f"\n[{i+1}/{len(clients)}] Processing: {client}")

        try:
            email_ids = search_emails_for_client(server, client, since_date)
            print(f"  Found {len(email_ids)} emails")

            if not email_ids:
                results.append({
                    "client_name": client,
                    "emails": [],
                })
                continue

            emails = fetch_emails(server, email_ids, max_emails=100)
            print(f"  Fetched {len(emails)} email details")

            results.append({
                "client_name": client,
                "emails": emails,
            })

            time.sleep(0.5)

        except Exception as e:
            print(f"  [ERROR] {e}")
            results.append({
                "client_name": client,
                "emails": [],
            })

    # 4. Disconnect
    server.logout()

    # 6. Upload to Supabase
    total_fetched_emails = upload_to_supabase(results)

    # 7. Summary
    print("\n" + "=" * 60)
    print("DONE! Summary:")
    print(f"  Total clients: {len(clients)}")
    print(f"  Total emails fetched and uploaded: {total_fetched_emails}")
    print("=" * 60)


if __name__ == "__main__":
    main()
